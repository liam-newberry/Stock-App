# File created by: Liam Newberry
import openpyxl
import xlsxwriter
import pandas as pd
from datetime import datetime, timedelta
import os
from settings import *

def get_tickers():
    wb_obj = openpyxl.load_workbook("portfolio_history.xlsx")
    sheet_obj = wb_obj.active
    tickers = []
    for i in range(2,52):
        cell_obj = sheet_obj.cell(row=1, column=i)
        if cell_obj.value != None:
            tickers.append(cell_obj.value)
    return tickers

def create_xlsx(name:str):
    wb = xlsxwriter.Workbook(name)
    ws = wb.add_worksheet()
    wb.close()

def date_to_list(date:str):
    date = str(date)
    date = date[:date.index(" ")]
    year = int(date[:date.index("-")])
    date = date[date.index("-") + 1:]
    month = int(date[:date.index("-")])
    date = date[date.index("-") + 1:]
    day = int(date)
    return [year,month,day]

def check_column(path:str,abbr:str):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    for i in range(2,52):
        cell_obj = sheet_obj.cell(row=1, column=i)
        if cell_obj.value == abbr:
            return True
    return False

def get_1y(data_frame:pd,abbr:str):
    dates = data_frame["Date"]
    stock = data_frame[abbr]
    dates = list(dates)
    dates2 = []
    today = date_to_list(dates[-1])
    one_year = today.copy()
    one_year[0] -= 1
    
    for i in dates:
        dates2.append(date_to_list(i))
    dates2.append(one_year)
    dates2.sort()
    dates2 = dates2[dates2.index(one_year) + 1:]
    
    stock = stock[len(stock)-len(dates2):]
    stock = list(stock)

    return dates2, stock

def get_stock(abbr:str,time:str,temp:bool=False):
    path = "portfolio_history.xlsx"
    if temp:
        path = "temp.xlsx"

    if not os.path.exists(path):
        create_xlsx(path)
    
    if not check_column(path,abbr):
        return False
    
    data_frame = pd.read_excel(path,sheet_name="Sheet1")
    
    return TIME_DICT[time](data_frame,abbr)

TIME_DICT = {"1y":get_1y}

# print(get_stock("TSLA","1y"))
print(get_tickers())

